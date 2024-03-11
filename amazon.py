import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from urllib.parse import urlparse
import concurrent.futures
import time
import os

def get_product_data(url, max_retries=5):
    for attempt in range(max_retries):
        try:
            print(f"Fetching data from: {url}")
            # Make a request to Amazon
            response = requests.get(url)
            print(f"Response status code: {response.status_code}")

            if response.status_code == 200:
                # Parse the HTML content
                soup = BeautifulSoup(response.content, 'html.parser')

                # Extract product name
                isim_elem = soup.find('span', {'id': 'productTitle'})
                isim = isim_elem.get_text().strip() if isim_elem else "Ürün adı bulunamadı"

                # Extract product price
                fiyat_elem = soup.find('span', {'class': 'a-price-whole'})
                fiyat = fiyat_elem.get_text().strip() if fiyat_elem else "Ürün fiyatı bulunamadı"

                # Extract seller
                satici_elem = soup.find('a', {'id': 'sellerProfileTriggerId'})
                satici = satici_elem.get_text().strip() if satici_elem else "Satıcı bulunamadı"

                return isim, fiyat, satici
            elif response.status_code == 503:
                print(f"Server returned 503. Retrying in {2*attempt} seconds (attempt {attempt+1}/{max_retries})")
                time.sleep(2**attempt)  # Exponential backoff
            else:
                print(f"Unexpected response status code: {response.status_code}")
                break
        except Exception as e:
            print(f"An error occurred while processing {url}: {e}")
            break

    # If max retries reached or encounter unexpected error, return default values
    return "Bilgi alınamadı", "Bilgi alınamadı", "Bilgi alınamadı"

workbook = Workbook()
sheet = workbook.active

sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 40
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 30

header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
for col in range(1, 5):
    sheet.cell(row=1, column=col).fill = header_fill

sheet["A1"] = "Ürün Linki"
sheet["B1"] = "Ürün İsmi"
sheet["C1"] = "Ürün Fiyatı"
sheet["D1"] = "Satıcı"

with open("urun_linkleri.txt") as f:
    urun_linkleri = f.read().splitlines()

with concurrent.futures.ThreadPoolExecutor() as executor:
    results = executor.map(get_product_data, urun_linkleri)

    for index, (isim, fiyat, satici) in enumerate(results, start=2):
        current_url = urun_linkleri[index - 2]
        sheet[f"B{index}"] = isim
        sheet[f"C{index}"] = fiyat
        sheet[f"D{index}"] = satici
        sheet[f"A{index}"].hyperlink = current_url

workbook.save("urun_bilgileri.xlsx")

# Fixing the sheets1.xml issue
file_path = "urun_bilgileri.xlsx"
if os.path.exists(file_path):
    with open(file_path, "rb+") as file:
        content = file.read()
        content = content.replace(b'<sheet name="Sheet">', b'<sheet name="Sheet" sheetId="1" state="visible" r:id="rId1"/>')
        file.seek(0)
        file.write(content)
        file.truncate()

print("Done!")
